#!/usr/bin/env python3
"""De-identify student submissions before grading, and restore names afterward.

Implements the coded de-identification pattern FERPA contemplates at
34 CFR 99.31(b)(2): each student gets an opaque code that is *not* derived
from their name, NetID, or any other personal information, and the crosswalk
mapping codes back to students never leaves this machine.

    anonymize.py mask   <submissions_dir> [--out DIR] [--roster FILE]
    anonymize.py unmask <feedback_dir>    --roster FILE [--scores FILE]

`mask` copies submissions into a working directory with student-identifying
filenames replaced by codes, scrubs names out of file formats that store
text, and writes the crosswalk. Grade the masked copies.

`unmask` renames the generated feedback documents back to real student
filenames and substitutes real names into scores.xlsx, ready for upload.

IMPORTANT LIMITS -- read the report `mask` prints:
  * Scanned PDFs and images cannot be scrubbed. A handwritten name or a
    photo of a signed page stays identifying no matter what the file is
    called. These are listed as UNSCRUBBABLE.
  * Scrubbing is best-effort text replacement. It catches a name typed into
    a spreadsheet cell or a notebook markdown cell; it will not catch a name
    embedded in a chart image or a screenshot pasted into a workbook.
De-identification reduces exposure. It does not by itself resolve whether
sending student work to a given service is permitted -- that is a question
for your institution's registrar or FERPA officer.
"""

import argparse
import json
import os
import re
import secrets
import shutil
import stat
import sys
import zipfile
from pathlib import Path

# last_first_netid_whatever.ext -- the Learning Suite download convention.
SUBMISSION_RE = re.compile(r"^(?P<last>[^_]+)_(?P<first>[^_]+)_(?P<netid>[^_]+)_(?P<rest>.+)$")

# Formats that are zipped XML and can have text rewritten in place.
OOXML = {".xlsx", ".xlsm", ".docx", ".pptx"}
# Formats that are JSON text.
JSON_LIKE = {".ipynb"}
# Formats whose content we cannot inspect or scrub.
OPAQUE = {".pdf", ".png", ".jpg", ".jpeg", ".heic", ".zip", ".vsd", ".vsdx"}

SKIP_NAMES = {"Icon\r", "Icon", ".DS_Store", "Thumbs.db"}


def skip(p):
    return (p.name in SKIP_NAMES or p.name.startswith("~$")
            or p.name.startswith("._") or p.name.startswith(".~lock"))


def new_code(used):
    """Opaque, random, not derived from any student attribute."""
    while True:
        code = "S-" + secrets.token_hex(3).upper()
        if code not in used:
            used.add(code)
            return code


def name_variants(last, first, netid):
    """Strings worth scrubbing from file contents, longest first."""
    out = {
        f"{first} {last}", f"{last}, {first}", f"{first}{last}",
        f"{last} {first}", last, first, netid,
    }
    return sorted({v for v in out if len(v) >= 3}, key=len, reverse=True)


def compile_variants(variants):
    """Word-boundary patterns, longest first.

    Bare substring matching is not safe here: 'Nora' occurs inside the OOXML
    attribute name 'mc:Ignorable', and replacing it produces invalid XML that
    Excel refuses to open. Anchor every variant to word boundaries.
    """
    out = []
    for v in variants:
        lead = r"\b" if v[0].isalnum() else ""
        tail = r"\b" if v[-1].isalnum() else ""
        out.append(re.compile(lead + re.escape(v) + tail, re.IGNORECASE))
    return out


def scrub_text(text, patterns, code):
    hits = 0
    for pat in patterns:
        text, n = pat.subn(code, text)
        hits += n
    return text, hits


# XML text nodes only -- everything between '>' and the next '<'. Never
# touches tag names or attribute values, which is where the corruption risk is.
TEXT_NODE = re.compile(r">([^<>]+)<")


def scrub_xml_part(xml, patterns, code):
    hits = 0

    def repl(m):
        nonlocal hits
        scrubbed, n = scrub_text(m.group(1), patterns, code)
        hits += n
        return f">{scrubbed}<"

    return TEXT_NODE.sub(repl, xml), hits


def scrub_ooxml(path, variants, code):
    """Rewrite an OOXML package, scrubbing identifying strings from text nodes.

    Covers cell text (sharedStrings.xml), document body text, and the
    creator / lastModifiedBy fields in docProps -- all of which are text
    nodes. Markup is left untouched.
    """
    patterns = compile_variants(variants)
    tmp = path.with_suffix(path.suffix + ".tmp")
    hits = 0
    try:
        with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.endswith(".xml"):
                    try:
                        text = data.decode("utf-8")
                        text, n = scrub_xml_part(text, patterns, code)
                        hits += n
                        data = text.encode("utf-8")
                    except UnicodeDecodeError:
                        pass
                zout.writestr(item, data)
        tmp.replace(path)
        return hits, None
    except Exception as exc:
        tmp.unlink(missing_ok=True)
        return 0, str(exc)


CODE_RE = re.compile(r"S-[0-9A-F]{6}")


def restore_ooxml_text(path, by_code):
    """Replace masked codes with real names inside an OOXML package's text.

    unmask renames feedback files, but a document that shows "Student: S-7F3A2B"
    in its body still reaches the student that way. This rewrites text nodes
    only -- the same narrow surface scrub_ooxml touches -- so formatting,
    formulas, and cached values all survive untouched.
    """
    tmp = path.with_suffix(path.suffix + ".tmp")
    hits = 0

    def repl(m):
        nonlocal hits
        def sub_code(cm):
            nonlocal hits
            st = by_code.get(cm.group(0))
            if not st:
                return cm.group(0)
            hits += 1
            return f"{st['first']} {st['last']} ({st['netid']})"
        return ">" + CODE_RE.sub(sub_code, m.group(1)) + "<"

    try:
        with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.endswith(".xml"):
                    try:
                        data = TEXT_NODE.sub(repl, data.decode("utf-8")).encode("utf-8")
                    except UnicodeDecodeError:
                        pass
                zout.writestr(item, data)
        tmp.replace(path)
        return hits, None
    except Exception as exc:
        tmp.unlink(missing_ok=True)
        return 0, str(exc)


def header_columns(ws):
    """Locate the First Name / Last Name / Net ID columns in a scores sheet.

    Returns (header_row, {field: column_index}) or (None, {}) if the sheet
    has no recognizable header row.
    """
    wanted = {"firstname": "first", "lastname": "last", "netid": "netid"}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10)):
        found = {}
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            key = re.sub(r"[^a-z]", "", cell.value.lower())
            if key in wanted:
                found[wanted[key]] = cell.column
        if "netid" in found:
            return row[0].row, found
    return None, {}


def scrub_json(path, variants, code):
    """Scrub a notebook by walking parsed JSON, touching values but never keys."""
    patterns = compile_variants(variants)
    hits = 0

    def walk(node):
        nonlocal hits
        if isinstance(node, str):
            out, n = scrub_text(node, patterns, code)
            hits += n
            return out
        if isinstance(node, list):
            return [walk(x) for x in node]
        if isinstance(node, dict):
            return {k: walk(v) for k, v in node.items()}
        return node

    try:
        doc = json.loads(path.read_text(encoding="utf-8"))
        path.write_text(json.dumps(walk(doc), indent=1), encoding="utf-8")
        return hits, None
    except Exception as exc:
        return 0, str(exc)


def mask(args):
    src = args.submissions_dir.expanduser().resolve()
    if not src.is_dir():
        sys.exit(f"error: no such directory: {src}")

    out = (args.out or src.parent / "masked").expanduser().resolve()
    roster_path = (args.roster or src.parent / "roster.json").expanduser().resolve()

    if out.exists() and any(out.iterdir()):
        sys.exit(f"error: {out} exists and is not empty -- remove it or pass --out")
    out.mkdir(parents=True, exist_ok=True)

    files = [f for f in sorted(src.iterdir()) if f.is_file() and not skip(f)]
    if not files:
        sys.exit(f"error: no submissions found in {src}")

    students, used, unparsed = {}, set(), []
    unscrubbable, errors = [], []

    for f in files:
        m = SUBMISSION_RE.match(f.stem)
        if not m:
            unparsed.append(f.name)
            continue
        last, first, netid = m["last"], m["first"], m["netid"]
        key = netid.lower()
        if key not in students:
            students[key] = {
                "code": new_code(used), "last": last, "first": first,
                "netid": netid, "files": [],
            }
        st = students[key]
        dest_name = f"{st['code']}_{m['rest']}{f.suffix}"
        dest = out / dest_name
        shutil.copy2(f, dest)
        st["files"].append({"original": f.name, "masked": dest_name})

        variants = name_variants(last, first, netid)
        ext = f.suffix.lower()
        if ext in OOXML:
            hits, err = scrub_ooxml(dest, variants, st["code"])
            if err:
                errors.append((dest_name, err))
        elif ext in JSON_LIKE:
            hits, err = scrub_json(dest, variants, st["code"])
            if err:
                errors.append((dest_name, err))
        elif ext in OPAQUE:
            unscrubbable.append(dest_name)
        else:
            unscrubbable.append(dest_name)

    roster = {
        "source": str(src),
        "masked": str(out),
        "students": list(students.values()),
    }
    roster_path.write_text(json.dumps(roster, indent=2))
    os.chmod(roster_path, stat.S_IRUSR | stat.S_IWUSR)  # 600 -- owner only

    print(f"\nMasked {len(files) - len(unparsed)} file(s) for {len(students)} student(s)")
    print(f"  masked copies : {out}")
    print(f"  crosswalk     : {roster_path}  (mode 600)\n")

    if unparsed:
        print(f"  COULD NOT PARSE ({len(unparsed)}) -- not copied, handle by hand:")
        for n in unparsed:
            print(f"      {n}")
        print()
    if unscrubbable:
        print(f"  UNSCRUBBABLE ({len(unscrubbable)}) -- filename is masked, but content")
        print(f"  may still identify the student (handwriting, scans, embedded images):")
        for n in unscrubbable:
            print(f"      {n}")
        print()
    if errors:
        print(f"  ERRORS ({len(errors)}):")
        for n, e in errors:
            print(f"      {n}: {e}")
        print()

    print("  Grade the masked copies. Do not commit the crosswalk to git.\n")
    return 1 if (unparsed or errors) else 0


def unmask(args):
    roster_path = args.roster.expanduser().resolve()
    if not roster_path.exists():
        sys.exit(f"error: no crosswalk at {roster_path}")
    roster = json.loads(roster_path.read_text())

    by_code = {s["code"]: s for s in roster["students"]}
    fb = args.feedback_dir.expanduser().resolve()
    if not fb.is_dir():
        sys.exit(f"error: no such directory: {fb}")

    renamed, orphans, restored_bodies = 0, [], 0
    for f in sorted(fb.iterdir()):
        if not f.is_file() or skip(f):
            continue
        m = re.match(r"^(S-[0-9A-F]{6})_(.*)$", f.name)
        if not m:
            orphans.append(f.name)
            continue
        code, rest = m.group(1), m.group(2)
        st = by_code.get(code)
        if not st:
            orphans.append(f.name)
            continue
        new_name = f"{st['last']}_{st['first']}_{st['netid']}_{rest}"
        f = f.rename(f.with_name(new_name))
        renamed += 1
        # The code is in the document body too, not just the filename.
        if f.suffix.lower() in OOXML:
            hits, err = restore_ooxml_text(f, by_code)
            if err:
                print(f"warning: could not rewrite text in {f.name}: {err}")
            elif hits:
                restored_bodies += 1

    print(f"\nRestored {renamed} feedback file(s) in {fb}")
    if restored_bodies:
        print(f"  and replaced the code inside {restored_bodies} document body/bodies")

    if args.scores:
        scores = args.scores.expanduser().resolve()
        try:
            import openpyxl
            wb = openpyxl.load_workbook(scores)
            replaced, named = 0, 0
            for ws in wb.worksheets:
                _, cols = header_columns(ws)
                for row in ws.iter_rows():
                    for cell in row:
                        if not (isinstance(cell.value, str) and cell.value.strip() in by_code):
                            continue
                        st = by_code[cell.value.strip()]
                        in_netid_col = cols.get("netid") == cell.column
                        cell.value = st["netid"]
                        replaced += 1
                        # Only fill names alongside a code that sat in the Net ID
                        # column -- a code elsewhere is not a student row.
                        if in_netid_col:
                            for field in ("first", "last"):
                                col = cols.get(field)
                                if col:
                                    ws.cell(row=cell.row, column=col, value=st[field])
                                    named += 1
            wb.save(scores)
            msg = f"Replaced {replaced} code(s) with NetIDs in {scores.name}"
            if named:
                msg += f", and filled {named} name cell(s)"
            print(msg)

            # openpyxl writes formulas without cached values, so the totals and
            # statistics in scores.xlsx would read as blank until something
            # recalculates them. Do it here rather than leaving a silent trap.
            try:
                from recalc import try_recalc
                ok, err = try_recalc(scores)
                print("Recalculated formulas in " + scores.name if ok
                      else f"warning: {err}\n         run recalc.py on {scores.name} "
                           f"before uploading, or the totals will read as blank")
            except ImportError:
                print(f"warning: could not import recalc.py -- run it on {scores.name} "
                      "before uploading, or the totals will read as blank")
        except ImportError:
            print("warning: openpyxl not installed -- scores.xlsx left unmodified")
        except Exception as exc:
            print(f"warning: could not rewrite {scores.name}: {exc}")

    if orphans:
        print(f"\n  UNMATCHED ({len(orphans)}) -- left as-is:")
        for n in orphans:
            print(f"      {n}")
    print()
    return 1 if orphans else 0


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)

    m = sub.add_parser("mask", help="de-identify submissions before grading")
    m.add_argument("submissions_dir", type=Path)
    m.add_argument("--out", type=Path, help="where masked copies go (default: ../masked)")
    m.add_argument("--roster", type=Path, help="crosswalk path (default: ../roster.json)")
    m.set_defaults(func=mask)

    u = sub.add_parser("unmask", help="restore real names to generated feedback")
    u.add_argument("feedback_dir", type=Path)
    u.add_argument("--roster", type=Path, required=True)
    u.add_argument("--scores", type=Path, help="scores.xlsx to rewrite in place")
    u.set_defaults(func=unmask)

    args = ap.parse_args()
    sys.exit(args.func(args))


if __name__ == "__main__":
    main()
