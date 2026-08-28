#!/usr/bin/env python3
"""Export scores.xlsx as the CSV Learning Suite's grade import accepts.

Usage:
    export_upload_csv.py <scores.xlsx> [--assignment TITLE] [--out FILE]

Learning Suite wants a minimal sheet: "Net ID" in the top-left cell, the
assignment title in the cell to its right, and one row per student. Both
columns must have a header or the scores do not enter. Everything else in
scores.xlsx -- the max-points row, the per-rubric-item columns, the average
rows -- has to stay out, since an "Average (points)" row would import as a
student who does not exist.

Scores are written as plain numbers. Learning Suite accepts General, Number,
Percent, or Text and rejects anything else, so no formatting is applied.

Reads the *cached* value of each Total formula, which means scores.xlsx must
have been recalculated first. anonymize.py unmask does that and then calls
this; run it again by hand after adjusting a score, since the CSV is a
snapshot rather than a live view.
"""

import argparse
import csv
import re
import sys
from pathlib import Path

DEFAULT_NAME = "scores_upload.csv"

# Rows that carry statistics or rubric maxima rather than a student.
NON_STUDENT = re.compile(r"^(max\s*points?|average|median|std|count|total)\b", re.I)


def _norm(s):
    return re.sub(r"[^a-z]", "", s.lower()) if isinstance(s, str) else ""


def find_columns(ws):
    """Locate the header row and the Net ID / Total columns."""
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10)):
        netid = total = None
        for cell in row:
            key = _norm(cell.value)
            if key == "netid":
                netid = cell.column
            elif key == "total":
                total = cell.column
        if netid and total:
            return row[0].row, netid, total
    return None, None, None


def export(scores_path, assignment=None, out=None):
    """Write the upload CSV. Returns (path, row_count, skipped_rows)."""
    import openpyxl

    scores_path = Path(scores_path).expanduser().resolve()
    if not scores_path.is_file():
        raise FileNotFoundError(f"no such file: {scores_path}")

    ws = openpyxl.load_workbook(scores_path, data_only=True).active
    header_row, netid_col, total_col = find_columns(ws)
    if not header_row:
        raise ValueError("could not find both a 'Net ID' and a 'Total' column "
                         f"in the first 10 rows of {scores_path.name}")

    if not assignment:
        # The title cell names the assignment; fall back to the file's folder.
        title = ws.cell(1, 1).value
        assignment = (str(title).strip() if title and header_row > 1
                      else scores_path.parent.name)

    rows, skipped, blank_totals = [], [], 0
    for r in range(header_row + 1, ws.max_row + 1):
        netid = ws.cell(r, netid_col).value
        score = ws.cell(r, total_col).value
        if netid is None or not str(netid).strip():
            continue
        netid = str(netid).strip()
        if NON_STUDENT.match(netid):
            continue
        # A Total written by openpyxl and never recalculated reads back as None,
        # not as the formula text -- the cached value simply is not there.
        if score is None:
            blank_totals += 1
            skipped.append((netid, "no cached total -- run recalc.py on the workbook"))
            continue
        if not isinstance(score, (int, float)):
            skipped.append((netid, f"non-numeric total: {score!r}"))
            continue
        # Keep integers integral so the CSV reads 29 rather than 29.0.
        rows.append([netid, int(score) if float(score).is_integer() else score])

    if not rows:
        if blank_totals:
            raise ValueError(
                f"{scores_path.name} has {blank_totals} student row(s) whose Total is a "
                "formula with no cached value. Recalculate it first:\n"
                f"  python recalc.py {scores_path}")
        raise ValueError(f"no student rows found in {scores_path.name}")

    out = Path(out).expanduser().resolve() if out else scores_path.parent / DEFAULT_NAME
    with open(out, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["Net ID", assignment])
        w.writerows(rows)
    return out, len(rows), skipped


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("scores", type=Path)
    ap.add_argument("--assignment", help="column header (default: the sheet's title cell)")
    ap.add_argument("--out", type=Path, help=f"output path (default: {DEFAULT_NAME} beside scores)")
    args = ap.parse_args()
    try:
        out, n, skipped = export(args.scores, args.assignment, args.out)
    except Exception as exc:
        sys.exit(f"error: {exc}")
    print(f"Wrote {out} -- {n} student row(s)")
    if skipped:
        print(f"  SKIPPED ({len(skipped)}):")
        for netid, why in skipped:
            print(f"      {netid}: {why}")
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
